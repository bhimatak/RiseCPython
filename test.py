'''from openpyxl.workbook import workbook
from openpyxl import load_workbook
wb = load_workbook('test.xlsx')
# ws = wb.active
ws = wb['testSh']

col_a = ws['A']
# ws['A3'] = "Bhima"
count = 1;
for cell in col_a:
    #print(f'{cell.value}\n')
    cell.value = "shankar"
    print(f'{cell.value}\n')
    count+=1
wb.save('test.xlsx')'''



with open('testTxt.txt') as f:
    print(f, len(f))
    i = 0 
    #while i < (len(f)):
        
    
    
    
    '''
    for line in len(f):
        key = ""
        tl = line.strip()
        
        if tl.startswith("Question"):
            print("Q"+ tl[9:])'''
            
            
            
            
            
            
            
            
            
'''
        if (tl == "") or ("cpp" == tl) or ("Copy code" in tl):
            continue
        
        
        if   ("d)" in tl) or ("c)" in tl) or ("b)" in tl) or ("a)" in tl) or ("What" in tl) or ("?" in tl) or ("(" in tl) or (")" in tl) or ("{" in tl) or (";" in tl) or ("}" in tl) or ("main()" in tl):
            print(tl)'''